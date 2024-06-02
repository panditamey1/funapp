import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os
# Define the lists
            # "Big": [22, 18, 29, 7, 28, 12, 35, 3, 26, 0, 32, 15, 19, 4, 21, 2, 25],
            # "Orph": [1, 20, 14, 31, 9, 17, 34, 6],
            # "Small": [27, 13, 36, 11, 30, 8, 23, 10, 5, 24, 16, 33],

big = [22, 18, 29, 7, 28, 12, 35, 3, 26, 0, 32, 15, 19, 4, 21, 2, 25]
orphan = [1, 20, 14, 31, 9, 17, 34, 6]
small = [27, 13, 36, 11, 30, 8, 23, 10, 5, 24, 16, 33]

# read all csv files from csv_files folder
for file in os.listdir('csv_files'):

    filename = os.path.join('csv_files', file)  
    # Read the CSV file into a DataFrame
    df = pd.read_csv(filename)

    # Save DataFrame to Excel
    df.to_excel("data.xlsx", index=False)

    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook("data.xlsx")
    sheet = workbook.active

    # Define the colors
    big_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red for big
    small_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green for small
    orphan_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  # Blue for orphan
    # there is only one column in the excel file
    # Iterate over the rows and apply colors based on the lists
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value in big:
                cell.fill = big_fill
            elif cell.value in small:
                cell.fill = small_fill
            elif cell.value in orphan:
                cell.fill = orphan_fill

    # Save the modified workbook
    filename = os.path.splitext(filename)[0]
    workbook.save(filename + ".xlsx")
